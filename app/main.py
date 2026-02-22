from fastapi import FastAPI, Request, HTTPException, BackgroundTasks, Form
from fastapi.responses import FileResponse, JSONResponse
import uvicorn
import os
import tempfile
import uuid
from dotenv import load_dotenv
from fastapi.middleware.cors import CORSMiddleware
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import openpyxl
import logging
from pydantic import BaseModel, Field
from typing import Literal, Optional, Any, Dict
import time
import concurrent.futures

import asyncio
import urllib.request
import json as _json

# Try modern GenAI SDK first, fall back to legacy
GENAI_SDK = None
genai_modern = None
genai_legacy = None
try:
    import google.genai as genai_modern
    GENAI_SDK = "modern"
except Exception:
    genai_modern = None
    try:
        import google.generativeai as genai_legacy
        GENAI_SDK = "legacy"
    except Exception:
        genai_legacy = None
        GENAI_SDK = None

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Email configuration — uses Resend HTTP API (works on Render free tier)
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "")
# Free Resend plan must use onboarding@resend.dev unless domain is verified
MAIL_FROM = os.getenv("MAIL_FROM", "onboarding@resend.dev")
MAIL_USERNAME = os.getenv("MAIL_USERNAME", "")   # destination / owner email
USE_EMAIL = bool(RESEND_API_KEY)

if USE_EMAIL:
    logger.info("Resend configurado — from: %s", MAIL_FROM)
else:
    logger.warning("Email NO configurado: falta RESEND_API_KEY")

# Google Gemini configuration
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
USE_GOOGLE = os.getenv("USE_GOOGLE", "false").lower() in ("1", "true", "yes")
if GOOGLE_API_KEY and USE_GOOGLE:
    try:
        if GENAI_SDK == "modern" and genai_modern is not None:
            # modern SDK often uses a Client object; we'll prefer that at call time
            logger.info("Usando SDK moderno google.genai para Gemini")
        elif GENAI_SDK == "legacy" and genai_legacy is not None:
            try:
                genai_legacy.configure(api_key=GOOGLE_API_KEY)
                logger.info("google.generativeai configurado con la clave proporcionada")
            except Exception:
                logger.exception("No se pudo configurar google.generativeai con la clave proporcionada")
        else:
            logger.warning("No se detectó ninguna SDK de GenAI instalada (ni modern ni legacy)")
    except Exception:
        logger.exception("Error al intentar configurar GenAI SDK")

app = FastAPI(title="backendKiros - Cotizaciones con IA")
# CORS: permitir peticiones desde el frontend en desarrollo
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory store for downloadable temp files {file_id: (path, media_type, filename)}
_download_store: Dict[str, tuple] = {}


class QuoteRequest(BaseModel):
    servicio: str = Field(..., min_length=1)
    horas: int = Field(1, ge=1, description="Número de horas, entero >= 1")
    complejidad: Literal["baja", "media", "alta"] = "media"
    cliente: Optional[str] = Field("Cliente")
    nombre: Optional[str] = Field(None)
    telefono: Optional[str] = Field(None)
    correo: Optional[str] = Field(None)
    fecha: Optional[str] = Field(None)
    condiciones: Optional[str] = Field(None)
    vigencia: Optional[str] = Field("15 días")


def calcular_precio(horas: int, complejidad: str) -> int:
    base = 1000
    if complejidad == "alta":
        return base + horas * 200
    if complejidad == "baja":
        return base + horas * 50
    return base + horas * 100


async def generar_proforma_ai(
    nombre: str,
    telefono: str,
    correo: str,
    servicio: str,
    horas: int,
    complejidad: str,
    precio: int,
    fecha: str = None,
    condiciones: str = None,
    vigencia: str = "15 días"
) -> str:
    """
    Prompt optimizado para que el LLM genere una proforma
    que coincida con los datos del formulario del cliente.
    """
    prompt = (
        f"Genera una proforma profesional en formato de cotización oficial.\n\n"
        f"Encabezado:\n"
        f"- Empresa emisora: [Nombre de la empresa]\n"
        f"- Fecha: {fecha if fecha else '[Fecha actual]'}\n"
        f"- Número de Proforma: PRF-2026-XXXX\n\n"
        f"Datos del cliente:\n"
        f"- Nombre completo: {nombre}\n"
        f"- Teléfono: {telefono}\n"
        f"- Correo electrónico: {correo}\n\n"
        f"Detalles del servicio solicitado:\n"
        f"- Servicio: {servicio}\n"
        f"- Horas estimadas: {horas}\n"
        f"- Complejidad: {complejidad}\n"
        f"- Precio total: ${precio}\n\n"
        f"Instrucciones para el formato:\n"
        f"1. Organiza la información en secciones: Encabezado, Datos del Cliente, Detalles del Servicio, Resumen de Costos, Condiciones, Nota Final.\n"
        f"2. Usa un tono formal y claro.\n"
        f"3. Incluye una breve descripción del alcance del servicio.\n"
        f"4. Añade condiciones: {condiciones if condiciones else '[Condiciones de pago y entrega]'}.\n"
        f"5. Indica la vigencia de la oferta: {vigencia}.\n"
        f"6. Finaliza con la nota: 'Esta es una cotización referencial, sujeta a ajustes.'\n"
        f"7. Formatea la salida como documento listo para enviar por correo y exportar a PDF/Excel."
    )

    if GOOGLE_API_KEY and USE_GOOGLE:
        max_retries = int(os.getenv("GENAI_MAX_RETRIES", "3"))
        timeout_sec = float(os.getenv("GENAI_TIMEOUT", "10.0"))
        backoff_base = float(os.getenv("GENAI_BACKOFF_BASE", "1.0"))

        last_exc: Optional[Exception] = None
        for attempt in range(1, max_retries + 1):
            try:
                logger.info("Intentando generar proforma con GenAI (sdk=%s) intento %d/%d", GENAI_SDK, attempt, max_retries)

                def _call_genai() -> Any:
                    # Modern SDK: google.genai
                    if GENAI_SDK == "modern" and genai_modern is not None:
                        client_gen = genai_modern.Client(api_key=GOOGLE_API_KEY)
                        model_name = os.getenv("GOOGLE_MODEL", "gemini-2.0-flash")
                        return client_gen.models.generate_content(model=model_name, contents=prompt)

                    # Legacy SDK: google.generativeai
                    if GENAI_SDK == "legacy" and genai_legacy is not None:
                        model_obj = genai_legacy.GenerativeModel(os.getenv("GOOGLE_MODEL", "gemini-2.0-flash"))
                        return model_obj.generate_content(prompt)

                    raise RuntimeError("No hay SDK GenAI disponible para realizar la llamada")

                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    fut = ex.submit(_call_genai)
                    resp = fut.result(timeout=timeout_sec)

                # extraer texto de la respuesta de forma robusta
                text = None
                if isinstance(resp, dict):
                    candidates = resp.get("candidates") or resp.get("outputs") or []
                    if candidates and isinstance(candidates, list) and candidates[0]:
                        cand = candidates[0]
                        if isinstance(cand, dict):
                            text = cand.get("content") or cand.get("text") or cand.get("output")
                    text = text or resp.get("output_text") or resp.get("text") or resp.get("content")
                else:
                    # modern/legacy SDK: .text is the simplest accessor
                    text = getattr(resp, "text", None) or getattr(resp, "output_text", None) or getattr(resp, "outputs", None) or getattr(resp, "content", None)

                if text:
                    logger.info("Proforma generada por GenAI (long=%d)", len(str(text)))
                    return str(text).strip()
                else:
                    logger.warning("GenAI devolvió respuesta sin texto útil (intento %d): %r", attempt, resp)
                    last_exc = RuntimeError("Respuesta vacía de GenAI")
            except concurrent.futures.TimeoutError:
                last_exc = TimeoutError(f"Timeout al llamar a GenAI después de {timeout_sec}s")
                logger.warning("Timeout en llamada a GenAI (intento %d/%d)", attempt, max_retries)
            except Exception as e:
                last_exc = e
                logger.exception("Error en intento de GenAI (intento %d/%d): %s", attempt, max_retries, e)

            # backoff exponencial antes del siguiente intento
            if attempt < max_retries:
                sleep_time = backoff_base * (2 ** (attempt - 1))
                logger.info("Esperando %.1fs antes del siguiente intento", sleep_time)
                time.sleep(sleep_time)

        # si llegamos aquí, todos los intentos fallaron
        logger.error("Todos los intentos a GenAI fallaron: última excepción: %s", last_exc)

    # Fallback local si Gemini no está configurado o falló
    return (
        f"Proforma para {nombre}: Servicio {servicio}, {horas} horas, "
        f"complejidad {complejidad}. Precio estimado: {precio} dólares.\n\n"
        "Gracias por su interés. Esta proforma fue generada localmente."
    )


def _cleanup_file(path: str) -> None:
    try:
        os.unlink(path)
        logger.info("Archivo temporal eliminado: %s", path)
    except Exception:
        logger.exception("No se pudo eliminar el archivo temporal: %s", path)


@app.post("/cotizar")
async def cotizar(payload: QuoteRequest):
    precio = calcular_precio(payload.horas, payload.complejidad)
    nombre = payload.nombre or payload.cliente or "Cliente"
    telefono = payload.telefono or "No especificado"
    correo = payload.correo or ""

    proforma = await generar_proforma_ai(
        nombre=nombre,
        telefono=telefono,
        correo=correo or "No especificado",
        servicio=payload.servicio,
        horas=payload.horas,
        complejidad=payload.complejidad,
        precio=precio,
        fecha=payload.fecha,
        condiciones=payload.condiciones,
        vigencia=payload.vigencia or "15 días",
    )

    # Enviar correo al cliente si hay dirección y credenciales configuradas
    if correo:
        cuerpo_email = (
            f"Hola {nombre},\n\n"
            f"Gracias por contactarnos. Aquí está tu cotización:\n\n"
            f"Servicio: {payload.servicio}\n"
            f"Teléfono: {telefono}\n"
            f"Horas estimadas: {payload.horas}\n"
            f"Complejidad: {payload.complejidad}\n"
            f"Precio estimado: ${precio}\n\n"
            f"--- PROFORMA ---\n\n"
            f"{proforma}\n\n"
            f"Atentamente,\nKairos Digital Lab"
        )
        await _enviar_correo(
            destinatario=correo,
            asunto="Tu cotización - Kairos Digital Lab",
            cuerpo=cuerpo_email,
            adjuntos=[],
        )

    return JSONResponse({
        "precio": precio,
        "proforma": proforma,
        "email_enviado": USE_EMAIL and bool(correo),
    })


@app.post("/cotizar_pdf")
async def cotizar_pdf(payload: QuoteRequest, background_tasks: BackgroundTasks):
    precio = calcular_precio(payload.horas, payload.complejidad)
    proforma = await generar_proforma_ai(
        nombre=payload.nombre or payload.cliente or "Cliente",
        telefono=payload.telefono or "No especificado",
        correo=payload.correo or "No especificado",
        servicio=payload.servicio,
        horas=payload.horas,
        complejidad=payload.complejidad,
        precio=precio,
        fecha=payload.fecha,
        condiciones=payload.condiciones,
        vigencia=payload.vigencia or "15 días",
    )

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    filename = tmp.name
    tmp.close()

    c = canvas.Canvas(filename)
    c.setFont("Helvetica", 12)
    c.drawString(100, 780, f"Cotización para: {payload.nombre or payload.cliente}")
    c.drawString(100, 760, f"Servicio: {payload.servicio}")
    c.drawString(100, 740, f"Horas: {payload.horas}")
    c.drawString(100, 720, f"Complejidad: {payload.complejidad}")
    c.drawString(100, 700, f"Precio estimado: {precio} dólares")
    c.drawString(100, 660, "Proforma:")

    textobject = c.beginText(100, 640)
    for line in proforma.splitlines():
        textobject.textLine(line)
    c.drawText(textobject)
    c.showPage()
    c.save()

    background_tasks.add_task(_cleanup_file, filename)
    return FileResponse(filename, media_type="application/pdf", filename="cotizacion.pdf")


@app.post("/cotizar_excel")
async def cotizar_excel(payload: QuoteRequest, background_tasks: BackgroundTasks):
    precio = calcular_precio(payload.horas, payload.complejidad)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"
    ws.append(["Cliente", "Teléfono", "Correo", "Servicio", "Horas", "Complejidad", "Precio"])
    ws.append([payload.nombre or payload.cliente, payload.telefono, payload.correo, payload.servicio, payload.horas, payload.complejidad, precio])

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    filename = tmp.name
    tmp.close()

    wb.save(filename)

    background_tasks.add_task(_cleanup_file, filename)
    return FileResponse(filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="cotizacion.xlsx")


# ── Helper: generate PDF using reportlab ──────────────────────────────────────
def _generar_pdf_proforma(contenido: str, nombre: str) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    filename = tmp.name
    tmp.close()

    page_w, page_h = A4
    c = canvas.Canvas(filename, pagesize=A4)
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, page_h - 50, "Proforma de Servicios - Kairos Digital Lab")
    c.setFont("Helvetica", 11)

    y = page_h - 80
    margin = 50
    line_h = 14

    for raw_line in contenido.splitlines():
        # wrap long lines
        words = raw_line.split()
        line = ""
        for word in words:
            test = (line + " " + word).strip()
            if c.stringWidth(test, "Helvetica", 11) < (page_w - 2 * margin):
                line = test
            else:
                c.drawString(margin, y, line)
                y -= line_h
                if y < 60:
                    c.showPage()
                    c.setFont("Helvetica", 11)
                    y = page_h - 50
                line = word
        if line:
            c.drawString(margin, y, line)
            y -= line_h
            if y < 60:
                c.showPage()
                c.setFont("Helvetica", 11)
                y = page_h - 50

    c.save()
    return filename


# ── Helper: generate Excel ────────────────────────────────────────────────────
def _generar_excel_proforma(contenido: str) -> str:
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    filename = tmp.name
    tmp.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proforma"
    for i, linea in enumerate(contenido.splitlines(), start=1):
        ws.cell(row=i, column=1, value=linea)
    wb.save(filename)
    return filename


# ── Helper: send email via Resend HTTP API (works on Render free tier) ────────
def _resend_send(destinatario: str, asunto: str, cuerpo: str) -> dict:
    """Calls Resend API over HTTPS — not blocked by Render."""
    payload = _json.dumps({
        "from": MAIL_FROM,
        "to": [destinatario],
        "subject": asunto,
        "text": cuerpo,
    }).encode("utf-8")
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=payload,
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            return _json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Resend HTTP {e.code}: {body}") from e


async def _enviar_correo(destinatario: str, asunto: str, cuerpo: str, adjuntos: list) -> None:
    if not USE_EMAIL:
        logger.warning("Email omitido: RESEND_API_KEY no configurado")
        return
    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, _resend_send, destinatario, asunto, cuerpo)
        logger.info("Correo enviado via Resend a %s — id=%s", destinatario, result.get("id"))
    except Exception:
        logger.exception("Error al enviar correo a %s", destinatario)


# ── GET /api/download/{file_id} ───────────────────────────────────────────────
@app.get("/api/download/{file_id}")
async def download_file(file_id: str, background_tasks: BackgroundTasks):
    entry = _download_store.pop(file_id, None)
    if not entry:
        raise HTTPException(status_code=404, detail="Archivo no encontrado o ya descargado")
    path, media_type, dl_name = entry
    background_tasks.add_task(_cleanup_file, path)
    return FileResponse(path, media_type=media_type, filename=dl_name)


# ── POST /api/proforma ────────────────────────────────────────────────────────
class ProformaRequest(BaseModel):
    nombre: str
    telefono: str
    correo: str
    servicio: str
    horas: int = Field(1, ge=1)
    complejidad: Literal["baja", "media", "alta"] = "media"
    precio: Optional[int] = None
    fecha: Optional[str] = None
    condiciones: Optional[str] = None
    vigencia: Optional[str] = "15 días"


@app.post("/api/proforma")
async def crear_proforma(payload: ProformaRequest):
    # 1. Calcular precio si no viene del cliente
    precio = payload.precio if payload.precio is not None else calcular_precio(payload.horas, payload.complejidad)

    # 2. Generar proforma con IA
    contenido = await generar_proforma_ai(
        nombre=payload.nombre,
        telefono=payload.telefono,
        correo=payload.correo,
        servicio=payload.servicio,
        horas=payload.horas,
        complejidad=payload.complejidad,
        precio=precio,
        fecha=payload.fecha,
        condiciones=payload.condiciones,
        vigencia=payload.vigencia or "15 días",
    )

    # 3. Crear PDF y Excel
    pdf_path = _generar_pdf_proforma(contenido, payload.nombre)
    excel_path = _generar_excel_proforma(contenido)

    # 4. Registrar archivos para descarga
    pdf_id = str(uuid.uuid4())
    excel_id = str(uuid.uuid4())
    _download_store[pdf_id] = (pdf_path, "application/pdf", "proforma.pdf")
    _download_store[excel_id] = (excel_path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "proforma.xlsx")

    # 5. Enviar por correo (si está configurado)
    await _enviar_correo(
        destinatario=payload.correo,
        asunto="Proforma de Servicios - Kairos Digital Lab",
        cuerpo=contenido,
        adjuntos=[pdf_path, excel_path],
    )

    # 6. Responder al frontend
    return JSONResponse({
        "status": "success",
        "precio": precio,
        "proforma": contenido,
        "email_enviado": USE_EMAIL,
        "message": "La proforma fue enviada al correo proporcionado." if USE_EMAIL else "Proforma generada. Configura MAIL_USERNAME/MAIL_PASSWORD para envío por correo.",
        "download_links": {
            "pdf": f"/api/download/{pdf_id}",
            "excel": f"/api/download/{excel_id}",
        },
    })


if __name__ == "__main__":
    uvicorn.run("app.main:app", host=os.getenv("HOST", "0.0.0.0"), port=int(os.getenv("PORT", 8000)), reload=True)


# ── GET /test-email ───────────────────────────────────────────────────────────
@app.get("/test-email")
async def test_email():
    """Debug: sends a test email via Resend to MAIL_USERNAME and returns result."""
    config_info = {
        "RESEND_API_KEY_length": len(RESEND_API_KEY),
        "MAIL_FROM": MAIL_FROM,
        "MAIL_USERNAME": MAIL_USERNAME or "(vacío)",
        "USE_EMAIL": USE_EMAIL,
    }
    if not USE_EMAIL:
        return JSONResponse({"ok": False, "error": "RESEND_API_KEY no configurado", "config": config_info})
    try:
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None, _resend_send,
            MAIL_USERNAME or "pame2394@gmail.com",
            "Test Resend - Kairos Digital Lab",
            "Este es un correo de prueba desde el backend de Kairos."
        )
        return JSONResponse({"ok": True, "resend_response": result, "config": config_info})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e), "config": config_info})
